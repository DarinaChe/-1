from openpyxl import load_workbook, Workbook
from win32com.client import Dispatch

path12 = input('Введите расположение файла "Новая ЗП общая": ')

def dispatch_surname(path3, q, rangeN):
    wbook = Workbook()
    ws = wbook.active
    wbook.save(path3)
    xl = Dispatch("Excel.Application")
    wb1 = xl.Workbooks.Open(Filename=path12)
    wb2 = xl.Workbooks.Open(Filename=path3)
    ws1 = wb1.Worksheets(1)
    ws1.Copy(Before=wb2.Worksheets(1))
    wb2.Close(SaveChanges=True)
    xl.Quit()
    wbook = load_workbook(path3)
    ws = wbook.active
    ws.move_range(rangeN, rows=q, cols=0, translate=True)
    ws.delete_rows(6, 21)
    wbook.save(path3)
    return "ок"

path122 =path12.partition('Новая ЗП общая.xlsx')[0]


dispatch_surname(f'{path122}{"Васильев.xlsx"}', -1,"A6:AD6")
dispatch_surname(f'{path122}{"Семенов_К.xlsx"}', -2, "A7:AD7")
dispatch_surname(f'{path122}{"Павлов.xlsx"}', -3,"A8:AD8" )
dispatch_surname(f'{path122}{"Еремеев.xlsx"}', -4, "A9:AD9")
dispatch_surname(f'{path122}{"Тараскин.xlsx"}', -5, "A10:AD10")
dispatch_surname(f'{path122}{"Романюк.xlsx"}', -6, "A11:AD11")
dispatch_surname(f'{path122}{"Семенов_А.xlsx"}', -9, "A14:AD14")
dispatch_surname(f'{path122}{"Федутенко.xlsx"}', -10, "A15:AD15")
dispatch_surname(f'{path122}{"Петрушин.xlsx"}', -11, "A16:AD16")
dispatch_surname(f'{path122}{"Пчелин.xlsx"}', -12, "A17:AD17")
dispatch_surname(f'{path122}{"Варламов.xlsx"}', -13, "A18:AD18")
dispatch_surname(f'{path122}{"Крестин.xlsx"}', -14, "A19:AD19")
dispatch_surname(f'{path122}{"Феоктистов.xlsx"}', -17, "A22:AD22")
dispatch_surname(f'{path122}{"Мурашов.xlsx"}', -18,"A23:AD23")
dispatch_surname(f'{path122}{"Мельник.xlsx"}', -19, "A24:AD24")
dispatch_surname(f'{path122}{"Маркин.xlsx"}', -20, "A25:AD25")