import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch


path1 = input('Введите расположение файла "Новая ЗП общая": ')
path2 = input('Введите расположение файла "Рабочие_часы": ')

path11 = path1.partition('Новая ЗП общая.xlsx')
part = path11[0]

# 'C:\\Users\\admin\\Desktop\\ProjectLU1\\Новая ЗП общая.xlsx'
# 'C:\\Users\\admin\\Desktop\\ProjectLU1\\Рабочие_часы.xlsx'

wh = load_workbook(path2, data_only=True)
ws_wh = wh.active

# создание новгого листа копированием предыдущего
xl = Dispatch("Excel.Application")
wb1 = xl.Workbooks.Open(Filename=path1)
ws1 = wb1.Worksheets(1)
ws1.Copy(Before=wb1.Worksheets(1))
wb1.Close(SaveChanges=True)
xl.Quit()

d_now = datetime.datetime.now()
month = d_now.strftime("%m")
if 2 <= int(month) <= 12:
    month_new = int(month) - 1
elif int(month) == 1:
    month_new = 12


months_map = {1: 'Январь', 2: 'Февраль', 3: 'Март', 4:'Апрель', 5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август', 9:'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}
def mapper(month_new):
    return months_map[month_new]

xxx = mapper(month_new)

data_now = f'{xxx} {d_now.strftime("%Y")}'

salary = load_workbook(path1)
ws_s_new = salary.active
ws_s_new.title = data_now


# перенос данных с одного файла на другой, отталкиваясь от столбца "Норма"
def otrabotka(a):
    for row in ws_wh.iter_rows():
        for cell in row:
            if cell.value == "Норма":
                recycling = f'{get_column_letter(cell.col_idx + 1)}{a}'
                fines = f'{get_column_letter(cell.col_idx + 2)}{a}'
                premium = f'{get_column_letter(cell.col_idx + 3)}{a}'
                check = f'{get_column_letter(cell.col_idx + 4)}{a}'
                vacation = f'{get_column_letter(cell.col_idx - 1)}{a}'
                time_ill = f'{get_column_letter(cell.col_idx - 2)}{a}'
                days_worked = f'{get_column_letter(cell.col_idx - 3)}{a}'
                car = f'{get_column_letter(cell.col_idx - 4)}{a}'
    return recycling, fines, premium, check, vacation, time_ill, days_worked, car


for row in ws_wh.iter_rows():
    for cell in row:
        if cell.value == "Норма":
            norma = get_column_letter(cell.col_idx)

def surname(x, y, z):
    ws_s_new[f'{"B"}{y}'].value = ws_wh[f'{norma}{2}'].value
    if ws_wh[f'{"A"}{x}'].value == ws_s_new[f'{"C"}{y}'].value:
        ws_s_new[f'{"N"}{y}'].value = ws_wh[otrabotka(x)[6]].value
        ws_s_new[f'{"O"}{y}'].value = ws_wh[otrabotka(x)[5]].value
        ws_s_new[f'{"P"}{y}'].value = ws_wh[otrabotka(x)[4]].value
        ws_s_new[f'{"S"}{y}'].value = ws_wh[otrabotka(x)[0]].value
        ws_s_new[f'{"Y"}{y}'].value = ws_wh[otrabotka(x)[3]].value
        ws_s_new[f'{"AB"}{y}'].value = ws_wh[otrabotka(x)[1]].value
        ws_s_new[f'{"AA"}{y}'].value = ws_wh[otrabotka(x)[2]].value
        ws_s_new[f'{"Q"}{y}'].value = ws_wh[otrabotka(x)[7]].value
    else:
        print("Ошибка в несовпадении сотрудника", z)
    return ws_s_new[f'{"N"}{y}'].value, ws_s_new[f'{"O"}{y}'].value, ws_s_new[f'{"P"}{y}'].value, ws_s_new[f'{"S"}{y}'].value, ws_s_new[f'{"Y"}{y}'].value, ws_s_new[f'{"AB"}{y}'].value, ws_s_new[f'{"AA"}{y}'].value, ws_s_new[f'{"Q"}{y}'].value


surname(3, 6, 'Васильев')
surname(4, 18, 'Варламов')
surname(5, 19, 'Крестин')
surname(6, 16, 'Петрушин')
surname(7, 17, 'Пчелин')
surname(8, 8, 'Павлов')
surname(9, 14, 'Семенов Александр')
surname(10, 7, 'Семенов Константин')
surname(11, 15, 'Федутенко')
surname(12, 9, 'Еремеев')
surname(13, 10, 'Тараскин')
surname(14, 11, 'Романюк')
surname(15, 22, 'Феоктистов')
surname(16, 23, 'Мурашов')
surname(17, 24, 'Мельник')
surname(18, 25, 'Маркин')

b = salary.save(path1)



